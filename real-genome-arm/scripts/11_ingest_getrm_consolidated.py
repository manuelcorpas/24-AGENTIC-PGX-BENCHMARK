#!/usr/bin/env python3
"""
Convert the GeT-RM Consolidated PGx and HLA Table into the caller-truth schema.

WHY THIS EXISTS
The consolidated table is a WIDE sheet, one row per Coriell sample and one
column per gene, published by CDC GeT-RM as a single spreadsheet covering 363
samples and 34 genes/loci (Scheinfeldt et al., J Mol Diagn 2025;27:457-464,
doi:10.1016/j.jmoldx.2025.02.008). The caller-truth evaluator
(08_caller_truth_eval.py) wants a LONG (sample, gene, diplotype) table. This
script is the bridge, so the only human step left is the download itself.

THE DOWNLOAD IS AUTOMATED (corrected 2026-07-27)
An earlier version of this docstring said cdc.gov returns HTTP 403 to
non-browser clients and the download was therefore necessarily manual. That was
wrong. cdc.gov 403s on requests missing a full browser header set, not on
automated clients as such; with Accept, Accept-Language, Sec-Fetch-* and
Upgrade-Insecure-Requests it returns 200 for both the page and the asset. See
10b_fetch_getrm_consolidated.sh, which pins the URL and records a SHA-256.

WHAT IT REFUSES TO DO
It does not invent, impute or carry forward a genotype. Blank cells, "Not
tested", "ND", "N/A" and free text are dropped rather than admitted, and a
table that yields no usable genotypes is an error rather than an empty truth
set. A truth set that silently loses rows produces a concordance number that
looks fine and means nothing.

USAGE
    python real-genome-arm/scripts/11_ingest_getrm_consolidated.py \
        --input real-genome-arm/getrm/GeT-RM_consolidated.xlsx \
        --out   real-genome-arm/getrm/getrm_consensus.tsv
    # optionally restrict to the benchmark's genes:
    #   --genes CYP2C19 CYP2D6 TPMT NUDT15 DPYD
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
GETRM = HERE.parent / "getrm"

# Headings the sample column has appeared under across GeT-RM releases.
SAMPLE_HEADINGS = ("coriell", "sample", "cell line", "dna", "nigms", "nhgri")

# Cells that mean "no genotype here". Matched case-insensitively after strip.
NON_GENOTYPE = {"", "-", "--", "nd", "n/a", "na", "none", "not tested",
                "not determined", "not available", "unknown", "nt"}

# Columns that carry provenance or data-availability flags, not genotypes.
METADATA_HINTS = ("reference", "bam", "fastq", "1000 genomes", "note", "comment",
                  "source", "study", "method", "population", "sex", "ethnicity")

# A diplotype is two alleles round a slash. An allele is a star allele (*4,
# *2xN), an HLA name (B*57:01) or a named haplotype (H1, Normal). Prose does not
# pass, and neither does a nucleotide genotype.
# Colons must be allowed after a star: GeT-RM reports HLA as *57:01:01, and
# four of the benchmark's lethal-class loci are HLA.
ALLELE = r"(?:\*[\w\.\:]+|[A-Za-z0-9][\w\*\.\:\-]*)"
DIPLOTYPE_RE = re.compile(rf"^{ALLELE}\s*/\s*{ALLELE}$")
SINGLE_RE = re.compile(rf"^{ALLELE}$")

# The consolidated sheet interleaves rsID-specific SNP columns among the
# star-allele columns, with values like "G/A". Those match the SHAPE of a
# diplotype without being one. Admitted, they would be compared against PyPGx
# star-allele calls, mismatch every time, and understate concordance. The truth
# set therefore admits only calls expressed in an allele vocabulary: something
# containing a star allele or an HLA-style colon.
NUCLEOTIDE_RE = re.compile(r"^[ACGTUN]+$", re.IGNORECASE)


def _is_allele_vocabulary(dip: str) -> bool:
    parts = [p.strip() for p in dip.split("/")]
    if any(NUCLEOTIDE_RE.match(p) for p in parts):
        return False
    return "*" in dip or ":" in dip


def find_header_row(rows: list[list[str]]) -> int:
    """Index of the header row.

    The published sheet opens with a blank row and puts the header on the
    second. Assuming row 0 is the header yields columns that are all empty
    strings, a table that parses without error, and zero genotypes: a silent
    empty truth set, which is the one outcome this pipeline must never produce.
    So find the header rather than assume its position.
    """
    for i, row in enumerate(rows[:20]):
        cells = [(c or "").strip() for c in row]
        if sum(1 for c in cells if c) < 2:
            continue
        if any(any(h in c.lower() for h in SAMPLE_HEADINGS) for c in cells):
            return i
    # No recognisable sample heading: fall back to the first row that has
    # content, and let find_sample_column raise a useful error.
    for i, row in enumerate(rows):
        if any((c or "").strip() for c in row):
            return i
    return 0


def read_table(path: Path) -> list[dict]:
    """Read the consolidated table from .xlsx, .xls or a CSV/TSV export."""
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm", ".xls"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise SystemExit(
                "reading .xlsx needs openpyxl (pip install openpyxl), or export "
                "the sheet to CSV and pass that instead")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = [[("" if c is None else str(c)) for c in r]
                for r in ws.iter_rows(values_only=True)]
        wb.close()
        if not rows:
            raise SystemExit(f"{path} has no rows")
        h = find_header_row(rows)
        header = [(c or "").strip() for c in rows[h]]
        return [dict(zip(header, r)) for r in rows[h + 1:]]

    delim = "\t" if suffix in (".tsv", ".tab") else ","
    with path.open(newline="") as fh:
        rows = [[(c or "") for c in r] for r in csv.reader(fh, delimiter=delim)]
    if not rows:
        raise SystemExit(f"{path} has no rows")
    h = find_header_row(rows)
    header = [c.strip() for c in rows[h]]
    return [dict(zip(header, r)) for r in rows[h + 1:]]


def find_sample_column(fieldnames) -> str:
    for name in fieldnames:
        if name and any(h in name.lower() for h in SAMPLE_HEADINGS):
            return name
    raise SystemExit(
        "could not identify the sample-identifier column in "
        f"{list(fieldnames)!r}. Rename it to 'Coriell #' or pass a CSV whose "
        "first column is the Coriell identifier.")


def is_metadata_column(name: str) -> bool:
    low = (name or "").lower()
    return any(h in low for h in METADATA_HINTS)


def clean_diplotype(value: str) -> str | None:
    """Return a normalised diplotype, or None if the cell is not a genotype."""
    v = (value or "").strip()
    if v.lower() in NON_GENOTYPE:
        return None
    # GeT-RM sometimes footnotes a call; drop trailing footnote markers.
    v = re.sub(r"\s*\[[^\]]*\]$", "", v).strip()
    v = re.sub(r"\s+", " ", v)
    if DIPLOTYPE_RE.match(v):
        a, b = (p.strip() for p in v.split("/", 1))
        dip = f"{a}/{b}"
        return dip if _is_allele_vocabulary(dip) else None
    if SINGLE_RE.match(v) and _is_allele_vocabulary(v):
        return v
    return None


def to_long(records: list[dict], genes: set[str] | None = None,
            report: bool = False):
    """Wide sheet to long (sample, gene, diplotype) rows.

    With report=True, also returns the columns that yielded no usable call, so
    that what was dropped is stated rather than silently discarded.
    """
    if not records:
        return ([], {}) if report else []
    fieldnames = list(records[0].keys())
    sample_col = find_sample_column(fieldnames)
    gene_cols = [c for c in fieldnames
                 if c and c != sample_col and not is_metadata_column(c)]

    out: list[dict] = []
    kept: set[str] = set()
    for rec in records:
        sample = (rec.get(sample_col) or "").strip()
        if not sample:
            continue
        for col in gene_cols:
            gene = re.split(r"[\n\r]", col)[0].strip()
            if not gene:
                continue
            if genes is not None and gene not in genes:
                continue
            dip = clean_diplotype(rec.get(col, ""))
            if dip is None:
                continue
            out.append({"sample": sample, "gene": gene, "diplotype": dip})
            kept.add(gene)

    if report:
        dropped = {}
        for col in gene_cols:
            gene = re.split(r"[\n\r]", col)[0].strip()
            if gene and gene not in kept:
                dropped[gene] = col
        return out, dropped
    return out


def build(src: Path, out: Path | None = None, genes: set[str] | None = None) -> list[dict]:
    rows = to_long(read_table(src), genes)
    if not rows:
        raise SystemExit(
            f"{src} yielded no usable genotypes. Refusing to write an empty "
            "truth set: the evaluator would report a vacuous 100% concordance.")
    if out is not None:
        out.parent.mkdir(parents=True, exist_ok=True)
        with out.open("w", newline="") as fh:
            w = csv.DictWriter(fh, fieldnames=["sample", "gene", "diplotype"],
                               delimiter="\t")
            w.writeheader()
            w.writerows(rows)
    return rows


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[1])
    ap.add_argument("--input", type=Path, required=True,
                    help="GeT-RM consolidated table (.xlsx or a CSV export)")
    ap.add_argument("--out", type=Path, default=GETRM / "getrm_consensus.tsv")
    ap.add_argument("--genes", nargs="*", default=None,
                    help="restrict to these gene columns")
    args = ap.parse_args(argv)

    if not args.input.exists():
        sys.stderr.write(
            f"no input at {args.input}\n"
            "The consolidated table is a manual download; see "
            "real-genome-arm/getrm/README.md\n")
        return 1

    rows, dropped = to_long(read_table(args.input),
                            set(args.genes) if args.genes else None, report=True)
    if not rows:
        sys.stderr.write(
            f"{args.input} yielded no usable genotypes. Refusing to write an "
            "empty truth set: the evaluator would report a vacuous 100%.\n")
        return 1
    args.out.parent.mkdir(parents=True, exist_ok=True)
    with args.out.open("w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=["sample", "gene", "diplotype"],
                           delimiter="\t")
        w.writeheader()
        w.writerows(rows)

    samples = len({r["sample"] for r in rows})
    genes = sorted({r["gene"] for r in rows})
    print(f"wrote {args.out}")
    print(f"  {len(rows)} genotypes over {samples} samples and {len(genes)} genes")
    print(f"  genes kept: {', '.join(genes)}")
    if dropped:
        # Stated, not silently discarded: a reader must be able to see what the
        # truth set does not cover.
        print(f"  columns yielding no star-allele call ({len(dropped)}), excluded:")
        for g in sorted(dropped):
            print(f"    {g}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
